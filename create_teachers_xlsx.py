from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()
sheet = wb.active
sheet.title = "O'qituvchilar"

# Teachers data
teachers = [
  ("Achilova Gulnoza Mirzakulovna", "achilova", "0987654321"),
  ("Achilova Rano Xamidovna", "achilova2", "0987654321"),
  ("Achilova Zuxra Xudayberdiyevna", "achilova3", "0987654321"),
  ("Akramova Sabina Ilxomjon qizi", "akramova", "0987654321"),
  ("Alijanov Sherzod Valijonovich", "alijanov", "0987654321"),
  ("Alikulov Odil Davronovich", "alikulov", "0987654321"),
  ("Alikulova Umida Nafasovna", "alikulova", "0987654321"),
  ("Arakulov Jahongir Xudoykul o'g'li", "arakulov", "0987654321"),
  ("Aripova Mavluda Kabilovna", "aripova", "0987654321"),
  ("Axmedova Gulandan Muxamedjanovna", "axmedova", "0987654321"),
  ("Azizov Ravshan A'zamjon o'g'li", "azizov", "0987654321"),
  ("Bektemirova Dilnoza Saidbotir qizi", "bektemirova", "0987654321"),
  ("Dexxomova Layloxon Solivaliyevna", "dexxomova", "0987654321"),
  ("Djurayeva Diana Abdulbayevna", "djurayeva", "0987654321"),
  ("Egamqulov Doniyor Umidjon o'g'li", "egamqulov", "0987654321"),
  ("Eshonkulova Ra'no Abduraxmanovna", "eshonkulova", "0987654321"),
  ("Fazilova Guzal Usmonjonovna", "fazilova", "0987654321"),
  ("Gayimnazirova Sabina Botir qizi", "gayimnazirova", "0987654321"),
  ("Hamidova O'g'iloy Mirzoqul qizi", "hamidova", "0987654321"),
  ("Ikromova Iroda Rustamjon qizi", "ikromova", "0987654321"),
  ("Jonzakova Dilfuza Abdulkasimovna", "jonzakova", "0987654321"),
  ("Kaxxarov Xusan Madaminovich", "kaxxarov", "0987654321"),
  ("Kuralov Sojida Absamatovna", "kuralov", "0987654321"),
  ("Mamadaminova Muxayyo Nematjonovna", "mamadaminova", "0987654321"),
  ("Mamirova Nazira Gaybullayevna", "mamirova", "0987654321"),
  ("Maxmudov Elbek Baxtiyor o'g'li", "maxmudov", "0987654321"),
  ("Mirzaqosimov Javlon Nodir o'g'li", "mirzaqosimov", "0987654321"),
  ("Musurmanova Dilnoza Furkatovna", "musurmanova", "0987654321"),
  ("Muxamedov Islombek Ilhom o'g'li", "muxamedov", "0987654321"),
  ("Ochilova Sohiba Alisher qizi", "ochilova", "0987654321"),
  ("Oripova Adolat Artikulovna", "oripova", "0987654321"),
  ("Pirnazarov Abdumunnab Abdullayevich", "pirnazarov", "0987654321"),
  ("Primova Zulfiya Shirinkulovna", "primova", "0987654321"),
  ("Qo'ylibayev Zuxriddin Karimjon o'g'li", "qoylibayev", "0987654321"),
  ("Qosimova Nigora Akmal qizi", "qosimova", "0987654321"),
  ("Radijapova Manzura Abduraximovna", "radijapova", "0987654321"),
  ("Rizoyeva Manzuraxon Xakimjon qizi", "rizoyeva", "0987654321"),
  ("Sabirova Natalya Nikolayevna", "sabirova", "0987654321"),
  ("Saidov Alisher Xolbekovich", "saidov", "0987654321"),
  ("Saidova Nesipgul", "saidova", "0987654321"),
  ("Saidova Umida Xolbekovna", "saidova2", "0987654321"),
  ("Saparova Zarema Abdukarimovna", "saparova", "0987654321"),
  ("Shamenova Saltanat Mukatirova qizi", "shamenova", "0987654321"),
  ("Sharikova Tatyana Vladimirovna", "sharikova", "0987654321"),
  ("Sherov Dilmurod Kuvandikovich", "sherov", "0987654321"),
  ("Tillavova Dianna Muxamedovna", "tillavova", "0987654321"),
  ("Tugalova Emine Ziyadinovna", "tugalova", "0987654321"),
  ("Tugalova Liliya Shevketovna", "tugalova2", "0987654321"),
  ("Turg'unboyeva Durdona Shuxrat qizi", "turgunboyeva", "0987654321"),
  ("Tuvalova Gulnoza Batirovna", "tuvalova", "0987654321"),
  ("Ubaydullayeva Sitora Xamidullo qizi", "ubaydullayeva", "0987654321"),
  ("Vaxidova Svetlana Samadovna", "vaxidova", "0987654321"),
  ("Xakimova Dilshoda Farqat qizi", "xakimova", "0987654321"),
  ("Xamrabayeva Lola Mamajon qizi", "xamrabayeva", "0987654321"),
  ("Xamroboyev Nozim Nasimjonovich", "xamroboyev", "0987654321"),
  ("Xaydarova Ramilya Muxamatshamirovna", "xaydarova", "0987654321"),
  ("Xershko Emine Fikriyevna", "xershko", "0987654321"),
  ("Xodjayeva Dildora Abdugaparovna", "xodjayeva", "0987654321"),
  ("Xudaykulova Umida Alibekovna", "xudaykulova", "0987654321"),
  ("Yulchiyev Baxtiyar Zokirjonovich", "yulchiyev", "0987654321"),
  ("Yuldasheva Xalkiz Kantureyevna", "yuldasheva", "0987654321"),
  ("Yusufov Shaxzod Shavkat o'g'li", "yusufov", "0987654321"),
  ("Zakirayeva Sharofat Safarovna", "zakirayeva", "0987654321"),
  ("Zuldayeva Elmira Usmanovna", "zuldayeva", "0987654321"),
]

# Styles
header_fill = PatternFill(start_color="1B3F66", end_color="1B3F66", fill_type="solid")
header_font = Font(name='Times New Roman', color="FFFFFF", bold=True, size=12)
title_font = Font(name='Times New Roman', size=18, bold=True)
border = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)

# Title
sheet['B2'] = "39-MAKTAB O'QITUVCHILARI LOGIN PAROL RO'YXATI"
sheet['B2'].font = title_font
sheet.merge_cells('B2:D2')

# Headers
headers = ["№", "F.I.O. (To'liq ism)", "Login", "Parol"]
for col, header in enumerate(headers, start=2):
    cell = sheet.cell(row=4, column=col)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border

# Data
for i, (name, login, password) in enumerate(teachers, start=1):
    row = i + 4
    
    # Number
    cell = sheet.cell(row=row, column=2)
    cell.value = i
    cell.alignment = Alignment(horizontal='center')
    cell.border = border
    
    # Name
    cell = sheet.cell(row=row, column=3)
    cell.value = name
    cell.border = border
    
    # Login
    cell = sheet.cell(row=row, column=4)
    cell.value = login
    cell.alignment = Alignment(horizontal='center')
    cell.border = border
    
    # Password
    cell = sheet.cell(row=row, column=5)
    cell.value = password
    cell.alignment = Alignment(horizontal='center')
    cell.border = border

# Column widths
sheet.column_dimensions['A'].width = 3
sheet.column_dimensions['B'].width = 6
sheet.column_dimensions['C'].width = 40
sheet.column_dimensions['D'].width = 18
sheet.column_dimensions['E'].width = 15

# Row heights
sheet.row_dimensions[2].height = 30
sheet.row_dimensions[4].height = 25

# Save
wb.save('/home/z/my-project/download/39_maktab_oqituvchilar_login_parol.xlsx')
print("Fayl yaratildi!")
